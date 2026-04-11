"""
DocXL AI — Excel Builder

Takes the result dict from pipeline.py and writes a formatted .xlsx file.

Styling:
  - Font: Arial throughout
  - Header rows: dark background (#1F3864), white bold text
  - Section labels: medium blue (#2E75B6), white bold text
  - Key column: light blue (#E8F4FD), dark blue text
  - Alternating data rows: #D6E4F0 and #FFFFFF
  - All cells: thin border (#B8CCE4)
  - Column widths: auto-fit, min 12, max 50
"""
import pandas as pd
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
    NamedStyle,
)
from openpyxl.utils import get_column_letter


# ── Colour constants ────────────────────────────────
_DARK_BG   = "1F3864"  # header rows
_MED_BLUE  = "2E75B6"  # section labels
_LIGHT_BG  = "E8F4FD"  # key column
_ALT_ROW_A = "D6E4F0"  # alternating row A
_ALT_ROW_B = "FFFFFF"  # alternating row B
_BORDER_CLR = "B8CCE4"
_DARK_TEXT  = "1F3864"
_WHITE      = "FFFFFF"

# ── Reusable style objects ──────────────────────────
_THIN_SIDE = Side(style="thin", color=_BORDER_CLR)
_ALL_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)
_FONT_ARIAL     = Font(name="Arial", size=10)
_FONT_HEADER    = Font(name="Arial", size=10, bold=True, color=_WHITE)
_FONT_SECTION   = Font(name="Arial", size=11, bold=True, color=_WHITE)
_FONT_KEY       = Font(name="Arial", size=10, color=_DARK_TEXT)
_FILL_HEADER    = PatternFill("solid", fgColor=_DARK_BG)
_FILL_SECTION   = PatternFill("solid", fgColor=_MED_BLUE)
_FILL_KEY       = PatternFill("solid", fgColor=_LIGHT_BG)
_FILL_ALT_A     = PatternFill("solid", fgColor=_ALT_ROW_A)
_FILL_ALT_B     = PatternFill("solid", fgColor=_ALT_ROW_B)
_ALIGN_LEFT     = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_w: int = 12, max_w: int = 50) -> None:
    """Auto-fit column widths to content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = min(max(max_len + 4, min_w), max_w)
        ws.column_dimensions[col_letter].width = width


def build_excel(result: Dict[str, Any], output_path: str) -> str:
    """
    Takes the result dict from pipeline.py and writes a formatted .xlsx file.

    Sheet structure:
      - If metadata dict is non-empty → Sheet 1 named 'Form Details'
        Columns: Field | Value
        Section headers use merged cells with coloured background
      - For each DataFrame in result['tables'] → one sheet named
        'Table 1', 'Table 2', etc.
        First row = column headers (bold, coloured background)
        Alternating row colours for data rows

    Returns output_path on success.
    """
    wb = Workbook()
    # Remove default sheet; we'll create our own
    if wb.active:
        wb.remove(wb.active)

    metadata: dict = result.get("metadata", {})
    tables: list = result.get("tables", [])
    source: str = result.get("source", "unknown")

    sheet_count = 0

    # ── Metadata sheet (Form Details) ────────────────────────
    if metadata:
        ws = wb.create_sheet(title="Form Details")
        sheet_count += 1

        # Section header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        hdr_cell = ws.cell(row=1, column=1, value="Extracted Form Details")
        hdr_cell.font = _FONT_SECTION
        hdr_cell.fill = _FILL_SECTION
        hdr_cell.alignment = _ALIGN_CENTER
        hdr_cell.border = _ALL_BORDER
        ws.cell(row=1, column=2).border = _ALL_BORDER

        # Column headers
        for ci, label in enumerate(["Field", "Value"], start=1):
            c = ws.cell(row=2, column=ci, value=label)
            c.font = _FONT_HEADER
            c.fill = _FILL_HEADER
            c.alignment = _ALIGN_CENTER
            c.border = _ALL_BORDER

        # Data rows
        for ri, (field, value) in enumerate(metadata.items(), start=3):
            # Key column
            kc = ws.cell(row=ri, column=1, value=field)
            kc.font = _FONT_KEY
            kc.fill = _FILL_KEY
            kc.alignment = _ALIGN_LEFT
            kc.border = _ALL_BORDER

            # Value column
            alt_fill = _FILL_ALT_A if (ri % 2 == 1) else _FILL_ALT_B
            vc = ws.cell(row=ri, column=2, value=str(value))
            vc.font = _FONT_ARIAL
            vc.fill = alt_fill
            vc.alignment = _ALIGN_LEFT
            vc.border = _ALL_BORDER

        _auto_width(ws)

    # ── Table sheets ─────────────────────────────────────────
    for idx, df in enumerate(tables):
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue
        sheet_count += 1
        sheet_name = f"Table {idx + 1}" if len(tables) > 1 else "Extracted Data"
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit

        # Header row
        for ci, col_name in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=ci, value=str(col_name))
            c.font = _FONT_HEADER
            c.fill = _FILL_HEADER
            c.alignment = _ALIGN_CENTER
            c.border = _ALL_BORDER

        # Data rows
        for ri, (_, row_data) in enumerate(df.iterrows(), start=2):
            alt_fill = _FILL_ALT_A if (ri % 2 == 0) else _FILL_ALT_B
            for ci, val in enumerate(row_data, start=1):
                c = ws.cell(row=ri, column=ci, value=val if pd.notna(val) else "")
                c.font = _FONT_ARIAL
                c.fill = alt_fill
                c.alignment = _ALIGN_LEFT
                c.border = _ALL_BORDER

        _auto_width(ws)

    # ── Safety: ensure at least one sheet exists ─────────────
    if sheet_count == 0:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value="No structured data could be extracted.")
        ws.cell(row=2, column=1, value=f"Extraction source: {source}")

    wb.save(output_path)
    return output_path
