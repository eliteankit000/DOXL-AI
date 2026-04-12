import tempfile
import re
import os
import uuid
import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════
# STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════

S_TITLE_BG    = "1F3864"   # dark navy   — title rows
S_SUBHEAD_BG  = "2E75B6"   # medium blue — section/subheading rows
S_HEADER_BG   = "FFC000"   # gold        — column header rows
S_KEY_BG      = "E8F4FD"   # light blue  — key column in KV sheets
S_ALT1_BG     = "D6E4F0"   # soft blue   — alternating odd rows
S_ALT2_BG     = "FFFFFF"   # white       — alternating even rows
S_TOTAL_BG    = "EAF3DE"   # light green — total/summary rows
S_WHITE_FG    = "FFFFFF"
S_DARK_FG     = "1F3864"
S_BLACK_FG    = "000000"


def _border():
    s = Side(border_style="thin", color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)


def _wc(ws, row, col, value,
        bold=False, size=10, bg=S_ALT2_BG,
        fg=S_BLACK_FG, align="left", wrap=False):
    """Write one cell with full formatting."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=size,
                       bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align,
                            vertical="center",
                            wrap_text=wrap)
    c.border    = _border()
    return c


def _merge_row(ws, row, num_cols, value,
               bold=True, size=11,
               bg=S_TITLE_BG, fg=S_WHITE_FG, height=26):
    """Write a full-width merged title or section row."""
    end_col = get_column_letter(num_cols)
    ws.merge_cells("A{}:{}{}".format(row, end_col, row))
    _wc(ws, row, 1, value, bold=bold, size=size,
        bg=bg, fg=fg, align="center")
    ws.row_dimensions[row].height = height


def _col_width(values: list, header: str) -> float:
    """Calculate best column width from content."""
    max_len = len(str(header))
    for v in values:
        max_len = max(max_len, len(str(v)) if v else 0)
    return min(max(max_len + 4, 10), 60)


# ═══════════════════════════════════════════════════════════
# DOCUMENT TYPE DETECTION
# ═══════════════════════════════════════════════════════════

# Document type constants
DOC_PURE_TABLE = "pure_table"   # inventory, attendance, stock
DOC_FORM       = "form"         # registration, admit card
DOC_MIXED_BILL = "mixed_bill"   # invoice, bill, receipt, room bill

BILL_SIGNALS = [
    r'\bbill\b', r'\binvoice\b', r'\breceipt\b',
    r'\bbill\s*no', r'\binvoice\s*no', r'\bgstn?o?\b',
    r'\bnet\s*amount\b', r'\bsub\s*total\b', r'\bpay\s*mode\b',
    r'\bcgst\b', r'\bsgst\b', r'\broom\s*bill\b',
    r'\btax\s*summary\b', r'\bparticulars?\b',
    r'\bwaiter\b', r'\btable\s*no\b', r'\bguest\b',
    r'\bcheckin\b', r'\bcheck.?in\b', r'\bcheck.?out\b',
]

TABLE_SIGNALS = [
    r'\bopening\s*stock\b', r'\bclosing\s*stock\b',
    r'\battendance\b', r'\bpresent\b', r'\babsent\b',
    r'\binventory\b', r'\bitem\s*name\b',
    r'\bpurchased\b', r'\bwastage\b', r'\bconsumption\b',
]


def detect_document_type(pdf_path: str) -> str:
    """
    Reads first page raw text and classifies the document.
    Returns one of: DOC_PURE_TABLE, DOC_FORM, DOC_MIXED_BILL
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return DOC_FORM
            text = (pdf.pages[0].extract_text(layout=False) or "").lower()
    except Exception:
        return DOC_FORM

    bill_score  = sum(1 for p in BILL_SIGNALS
                      if re.search(p, text, re.IGNORECASE))
    table_score = sum(1 for p in TABLE_SIGNALS
                      if re.search(p, text, re.IGNORECASE))

    if bill_score >= 2:
        return DOC_MIXED_BILL
    if table_score >= 2:
        return DOC_PURE_TABLE
    return DOC_FORM


# ═══════════════════════════════════════════════════════════
# BILL TITLE EXTRACTOR
# ═══════════════════════════════════════════════════════════

def _extract_bill_title(lines: list) -> tuple:
    """
    Returns (doc_title, doc_subtitle) for bill documents.

    Title priority order:
      1. A line that is a known bill type keyword
      2. The hotel/company name (short non-address line)
      3. Fallback: first non-address, non-date line

    Address lines are NEVER the title.
    """
    BILL_TYPE_KEYWORDS = [
        r'^BILL$', r'^INVOICE$', r'^RECEIPT$',
        r'^ROOM\s+BILL$', r'^TAX\s+INVOICE$',
        r'^PROFORMA\s+INVOICE$', r'^CREDIT\s+NOTE$',
        r'^DEBIT\s+NOTE$', r'^CASH\s+MEMO$',
        r'^DELIVERY\s+CHALLAN$', r'^QUOTATION$',
        r'^ESTIMATE$', r'^PURCHASE\s+ORDER$',
    ]

    ADDRESS_SIGNALS = [
        r'\d{6}',                    # pin code
        r'\b(road|marg|nagar|colony|street|lane|'
        r'sector|phase|plot|flat|near|behind|'
        r'opposite|opp\.)\b',
        r'\d{10,}',                  # phone number
        r'@',                        # email
        r'\b(mp|up|delhi|mumbai|'
        r'gujarat|maharashtra|'
        r'rajasthan)\b',
        r'^[0-9]',                   # starts with number
        r'[,].{1,5}[,]',            # "city, state, pin"
    ]

    def is_address(line):
        for p in ADDRESS_SIGNALS:
            if re.search(p, line, re.IGNORECASE):
                return True
        return False

    doc_title    = ""
    doc_subtitle = ""

    # Priority 1: Find bill type keyword line
    for line in lines[:20]:
        for kw in BILL_TYPE_KEYWORDS:
            if re.match(kw, line.strip(), re.IGNORECASE):
                doc_title = line.strip().upper()
                break
        if doc_title:
            break

    # Priority 2: Find company/hotel name
    company_name = ""
    for line in lines[:15]:
        line = line.strip()
        if (3 < len(line) <= 50
                and not is_address(line)
                and not re.match(
                    r'^\d{1,2}[/\-]\d{1,2}', line)
                and not re.match(r'^\d+$', line)
                and not re.match(r'^GSTNo', line,
                                 re.IGNORECASE)
                and not re.match(r'^Bill\s*No', line,
                                 re.IGNORECASE)):
            if not company_name:
                company_name = line

    if doc_title and company_name \
            and company_name != doc_title:
        doc_subtitle = company_name
    elif not doc_title and company_name:
        doc_title = company_name

    if not doc_title:
        doc_title = "Document"

    return doc_title, doc_subtitle


# ═══════════════════════════════════════════════════════════
# BILL METADATA EXTRACTOR
# ═══════════════════════════════════════════════════════════

def _extract_bill_metadata(raw_text: str,
                            page_num: int) -> list:
    """
    Extracts all key:value pairs from bill/invoice text.
    Returns list of (key, value) tuples.
    Only runs on page 1 (page_num == 0).
    """
    if page_num > 0:
        return []

    pairs = []
    text = raw_text

    # Pattern 1: "Key : Value" or "Key :- Value"
    for m in re.finditer(
        r'([A-Za-z][A-Za-z0-9 ./\-_()*]{1,40}?)'
        r'\s*:[-\s]+([^\n:]{1,80})',
        text
    ):
        key = m.group(1).strip().rstrip('.-: ')
        val = m.group(2).strip()
        if (key and val
                and len(key) >= 2
                and len(key) <= 45
                and val not in ('-', ':', '')
                and not re.match(r'^\s*$', val)):
            if val.count(':') < 3:
                pairs.append((key, val))

    # Pattern 2: Inline pairs on same line
    for line in text.split('\n'):
        segments = re.findall(
            r'([A-Za-z][A-Za-z0-9 ./\-_]{1,30}?)'
            r'\s*[:.]+\s*([A-Za-z0-9][^\t:]{1,40})',
            line
        )
        for key, val in segments:
            key = key.strip().rstrip('.-: ')
            val = val.strip()
            if (key and val
                    and 2 <= len(key) <= 40
                    and len(val) >= 1):
                pairs.append((key, val))

    # Pattern 3: Summary amounts
    SUMMARY_PATTERNS = [
        (r'Sub\s*Total\s*[:\-]+\s*([\d,]+\.?\d*)',
         'Sub Total'),
        (r'Net\s*Amount\s*[:\-]+\s*([\d,]+\.?\d*)',
         'Net Amount'),
        (r'Round\s*[Oo]ff\s*[:\-]+\s*([\d,.-]+)',
         'Round Off'),
        (r'Tax\s*Amount\s*[:\-]+\s*([\d,]+\.?\d*)',
         'Tax Amount'),
        (r'Balance\s*[:\-]+\s*([\d,]+\.?\d*)',
         'Balance'),
        (r'Received\s*Amount\s*[:\-]+\s*([\d,]+\.?\d*)',
         'Received Amount'),
        (r'Pay(?:ment)?\s*Mode\s*[:\-]+\s*(.+)',
         'Payment Mode'),
    ]
    for pattern, label in SUMMARY_PATTERNS:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            pairs.append((label, m.group(1).strip()))

    return pairs


# ═══════════════════════════════════════════════════════════
# PHANTOM COLUMN ELIMINATION
# ═══════════════════════════════════════════════════════════

def _remove_phantom_columns(headers: list,
                             data_rows: list) -> tuple:
    """
    Removes columns where:
      1. Header is Col_N (auto-generated, no real name)
      AND
      2. Less than 20% of data rows have a value in that column

    Returns (cleaned_headers, cleaned_rows)
    """
    if not headers or not data_rows:
        return headers, data_rows

    cols_to_keep = []
    for ci, h in enumerate(headers):
        # Always keep columns with real header names
        if not h.startswith("Col_"):
            cols_to_keep.append(ci)
            continue
        # For Col_N: keep only if data fill > 20%
        filled = sum(
            1 for row in data_rows
            if ci < len(row)
            and row[ci]
            and str(row[ci]).strip()
            and str(row[ci]).strip() != 'None'
        )
        if filled / max(len(data_rows), 1) > 0.20:
            cols_to_keep.append(ci)

    if not cols_to_keep:
        return headers, data_rows

    new_headers = [headers[i] for i in cols_to_keep]
    new_rows    = [
        [row[i] if i < len(row) else ""
         for i in cols_to_keep]
        for row in data_rows
    ]
    return new_headers, new_rows


# ═══════════════════════════════════════════════════════════
# EXISTING HELPERS (UNCHANGED)
# ═══════════════════════════════════════════════════════════

def _is_junk_row(row: list) -> bool:
    """Return True for rows that are headers/footers/noise."""
    non_empty = [str(c).strip() for c in row
                 if c and str(c).strip()]
    if not non_empty:
        return True

    joined = " ".join(non_empty).strip().lower()

    junk_patterns = [
        r"^\d{1,2}/\d{1,2}/\d{4}",
        r"^page\s*\d",
        r"http[s]?://",
        r"^-{3,}$",
        r"^_{3,}$",
        r"signature",
        r"^\s*\d+\s*$",
    ]
    for p in junk_patterns:
        if re.search(p, joined, re.IGNORECASE):
            return True

    if (len(non_empty) == 1
            and len(row) >= 4
            and len(non_empty[0]) > 25):
        return True

    if re.match(r"\d{1,2}/\d{1,2}/\d{4}\s+\d+:\d+", joined):
        return True

    if len(non_empty) / max(len(row), 1) < 0.2:
        return True

    return False


def _clean_cell(value) -> str:
    """Normalize a cell value to clean string."""
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s{2,}", " ", s)
    return s


def _find_header_row(table: list) -> int:
    """Find the index of the real header row."""
    for i, row in enumerate(table):
        non_empty = [c for c in row
                     if c and str(c).strip()
                     and not re.match(r"^[\d\s|.,-]+$",
                                      str(c).strip())]
        if len(non_empty) >= max(2, len(row) * 0.4):
            return i
    return 0


def _looks_like_data_row(row: list, known_headers: list) -> bool:
    """Returns True if this row looks like data, not a header."""
    if not known_headers:
        return False
    non_empty = [str(c).strip() for c in row
                 if c and str(c).strip()]
    if not non_empty:
        return True

    data_pattern = re.compile(
        r'\d+\s*(BOTTEL|BEER|ML|ml|L|PCS|KG|GM|Rs|INR|\|)',
        re.IGNORECASE
    )
    data_hits = sum(1 for v in non_empty
                    if data_pattern.search(v))
    if data_hits / len(non_empty) >= 0.4:
        return True

    numeric_hits = sum(1 for v in non_empty
                       if re.search(r'\d', v))
    if numeric_hits / len(non_empty) >= 0.5:
        return True

    header_set = set(h.lower().strip() for h in known_headers if h)
    cell_set = set(v.lower().strip() for v in non_empty if v)
    overlap = header_set & cell_set
    if overlap and len(overlap) >= len(header_set) * 0.5:
        return False

    return False


def _is_column_header_line(line: str,
                           all_table_headers: list) -> bool:
    """Returns True if this line is actually a column header row."""
    line_words = set(re.findall(r'[A-Z]{2,}', line.upper()))
    if len(line_words) < 3:
        return False
    for headers in all_table_headers:
        header_words = set()
        for h in headers:
            header_words.update(re.findall(r'[A-Z]{2,}',
                                           str(h).upper()))
        overlap = line_words & header_words
        if len(overlap) >= 3:
            return True
    return False


# ═══════════════════════════════════════════════════════════
# BILL / INVOICE EXTRACTOR (NEW)
# ═══════════════════════════════════════════════════════════

def extract_bill_content(pdf_path: str) -> dict:
    """
    Extracts bills, invoices, receipts, room bills.
    Uses ONLY geometric line strategy — NEVER text-alignment.
    Separates: header metadata, item tables, summary, footer.

    Returns same structure as extract_pdf_content():
    {
      "doc_title":       str,
      "doc_subtitle":    str,
      "tables":          [...],
      "metadata_pairs":  [...],
      "raw_text_pages":  [...]
    }
    """
    doc_title      = ""
    doc_subtitle   = ""
    all_tables     = []
    metadata_pairs = []
    raw_text_pages = []

    with pdfplumber.open(pdf_path) as pdf:

        for page_num, page in enumerate(pdf.pages):

            raw_text = page.extract_text(layout=False) or ""
            raw_text_pages.append(raw_text)
            lines = [l.strip() for l in raw_text.split('\n')
                     if l.strip()]

            # ── Extract title from first page only ──────────
            if page_num == 0:
                doc_title, doc_subtitle = \
                    _extract_bill_title(lines)

            # ── Extract metadata key-value pairs ────────────
            meta_from_page = _extract_bill_metadata(
                raw_text, page_num)
            metadata_pairs.extend(meta_from_page)

            # ── Extract tables using GEOMETRIC ONLY ─────────
            # NEVER use text-alignment on bills
            tables_raw = page.extract_tables({
                "vertical_strategy":   "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance":  3,
                "join_tolerance":  3,
                "edge_min_length": 3,
            }) or []

            for table in tables_raw:
                if not table or len(table) < 2:
                    continue

                hdr_idx = _find_header_row(table)
                raw_headers = table[hdr_idx]

                # Clean headers
                headers = []
                for i, h in enumerate(raw_headers):
                    val = _clean_cell(h)
                    headers.append(val if val else
                                   "Col_{}".format(i+1))

                # Remove trailing empty/Col_ columns that have
                # NO data in any row
                max_real_col = 0
                for row in table[hdr_idx+1:]:
                    for ci, cell in enumerate(row):
                        if cell and str(cell).strip():
                            max_real_col = max(max_real_col,
                                               ci + 1)
                headers = headers[:max_real_col]
                if not headers:
                    continue

                # Remove Col_N headers that have no header text
                # AND the column has sparse data (< 30% filled)
                real_data_rows = table[hdr_idx+1:]
                cols_to_keep = []
                for ci, h in enumerate(headers):
                    filled = sum(
                        1 for row in real_data_rows
                        if ci < len(row)
                        and row[ci]
                        and str(row[ci]).strip()
                    )
                    fill_rate = (filled /
                                 max(len(real_data_rows), 1))
                    if (not h.startswith("Col_")
                            or fill_rate > 0.25):
                        cols_to_keep.append(ci)

                if not cols_to_keep:
                    continue

                headers = [headers[i] for i in cols_to_keep]

                # Collect data rows
                data_rows = []
                for row in real_data_rows:
                    if _is_junk_row(row):
                        continue
                    clean = [_clean_cell(
                                 row[i] if i < len(row)
                                 else "")
                             for i in cols_to_keep]
                    if any(clean):
                        data_rows.append(clean)

                if not data_rows:
                    continue

                # Apply phantom column elimination
                headers, data_rows = _remove_phantom_columns(
                    headers, data_rows)
                if not headers or not data_rows:
                    continue

                # Sheet name from first column header
                sheet_name = headers[0][:28] if headers[0] \
                             else "Table {}".format(
                                 len(all_tables)+1)

                all_tables.append({
                    "heading":    sheet_name,
                    "headers":    headers,
                    "rows":       data_rows,
                    "page_start": page_num + 1,
                })

    # Deduplicate metadata pairs
    seen_keys = set()
    deduped_meta = []
    for k, v in metadata_pairs:
        if k not in seen_keys and v and v != '-':
            deduped_meta.append((k, v))
            seen_keys.add(k)

    return {
        "doc_title":      doc_title or "Document",
        "doc_subtitle":   doc_subtitle,
        "tables":         all_tables,
        "metadata_pairs": deduped_meta,
        "raw_text_pages": raw_text_pages,
    }


# ═══════════════════════════════════════════════════════════
# STEP 1 — EXTRACT ALL CONTENT FROM PDF (PURE TABLE / FORM)
# ═══════════════════════════════════════════════════════════

def extract_pdf_content(pdf_path: str) -> dict:
    """
    Extract all content from a PDF into a structured dict.
    Handles all document types: tables, forms, mixed, multi-page.
    (Used for DOC_PURE_TABLE and DOC_FORM — not for bills.)
    """
    doc_title      = ""
    doc_subtitle   = ""
    all_tables     = []
    metadata_pairs = []
    raw_text_pages = []

    with pdfplumber.open(pdf_path) as pdf:

        last_known_headers = []
        last_known_num_cols = 0

        for page_num, page in enumerate(pdf.pages):

            raw_text = page.extract_text(layout=False) or ""
            raw_text_pages.append(raw_text)
            lines = [ln.strip() for ln in raw_text.split("\n")
                     if ln.strip()]

            # ── Try geometric table extraction ──────────────
            tables_geo = page.extract_tables({
                "vertical_strategy":   "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance":      3,
                "join_tolerance":      3,
                "edge_min_length":     3,
            })

            if tables_geo:
                page_tables = tables_geo
            else:
                page_tables = page.extract_tables({
                    "vertical_strategy":   "text",
                    "horizontal_strategy": "text",
                    "snap_tolerance":      5,
                    "min_words_vertical":  1,
                }) or []

            # ── Process each table on this page ─────────────
            for table in page_tables:

                if not table or len(table) < 1:
                    continue

                hdr_idx = _find_header_row(table)
                raw_headers = table[hdr_idx]
                candidate_headers = [
                    _clean_cell(h) or "Col_{}".format(i+1)
                    for i, h in enumerate(raw_headers)
                ]

                is_continuation = (
                    last_known_headers
                    and len(candidate_headers) == last_known_num_cols
                    and _looks_like_data_row(raw_headers, last_known_headers)
                )

                if is_continuation:
                    headers = last_known_headers
                    start_from = 0
                else:
                    headers = candidate_headers
                    start_from = hdr_idx + 1
                    last_known_headers = headers
                    last_known_num_cols = len(headers)

                data_rows = []
                for row in table[start_from:]:
                    if _is_junk_row(row):
                        continue
                    non_empty_vals = [str(c).strip() for c in row
                                     if c and str(c).strip()]
                    if (len(non_empty_vals) == 1
                            and len(non_empty_vals[0]) > 20):
                        continue
                    clean = [_clean_cell(c) for c in row]
                    while len(clean) < len(headers):
                        clean.append("")
                    data_rows.append(clean[:len(headers)])

                if not data_rows:
                    continue

                # Apply phantom column elimination
                headers, data_rows = _remove_phantom_columns(
                    headers, data_rows)
                if not headers or not data_rows:
                    continue

                if is_continuation and all_tables:
                    all_tables[-1]["rows"].extend(data_rows)
                else:
                    all_tables.append({
                        "heading":    "",
                        "headers":    headers,
                        "rows":       data_rows,
                        "page_start": page_num + 1,
                    })

            # ── Document title from page 1 (AFTER tables) ────
            if page_num == 0:
                page_header_lists = []
                for t in (all_tables or []):
                    page_header_lists.append(t["headers"])

                title_candidates = []
                for line in lines:
                    line = line.strip()
                    if len(line) <= 8:
                        continue
                    if re.match(r"^\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}",
                                line):
                        continue
                    if re.match(r"^\d+$", line):
                        continue
                    if re.match(r"^\d{1,2}:\d{2}", line):
                        continue
                    if re.match(r"\d{1,2}/\d{1,2}/\d{4}\s+\d",
                                line):
                        continue
                    if "http" in line.lower():
                        continue
                    if _is_column_header_line(line,
                                             page_header_lists):
                        continue
                    title_candidates.append(line)
                    if len(title_candidates) == 3:
                        break

                if title_candidates:
                    doc_title = max(title_candidates, key=len)
                else:
                    doc_title = "Document"

                doc_subtitle = ""

            # ── Extract key-value metadata ───────────────────
            if not page_tables:
                for line in lines:
                    kv = re.match(
                        r"^([A-Za-z][A-Za-z0-9 './#&()\-]{1,50})"
                        r"\s*:\s*(.+)$",
                        line
                    )
                    if kv:
                        key = kv.group(1).strip().rstrip(":")
                        val = kv.group(2).strip()
                        if (key and val
                                and val not in ("-", ":", "")
                                and len(key) < 55):
                            metadata_pairs.append((key, val))

    return {
        "doc_title":      doc_title or "Document",
        "doc_subtitle":   doc_subtitle,
        "tables":         all_tables,
        "metadata_pairs": metadata_pairs,
        "raw_text_pages": raw_text_pages,
    }


# ═══════════════════════════════════════════════════════════
# STEP 2 — BUILD EXCEL FROM EXTRACTED CONTENT
# ═══════════════════════════════════════════════════════════

def build_excel(content: dict, output_path: str) -> str:
    """
    Build a fully formatted Excel file from extracted PDF content.
    Works for ANY document type. Zero hardcoded strings.
    """
    wb = Workbook()
    sheet_created = False

    # ── A: Write each table as its own sheet ──────────────────
    for t_idx, table in enumerate(content["tables"]):

        headers    = table["headers"]
        rows       = table["rows"]
        heading    = table["heading"]
        num_cols   = len(headers)

        if not rows:
            continue

        raw_name = headers[0] if headers[0] else \
                   "Table {}".format(t_idx + 1)
        for bad in ['\\', '/', ':', '*', '?', '[', ']']:
            raw_name = raw_name.replace(bad, " ")
        sheet_name = raw_name.strip()[:31] or \
                     "Table {}".format(t_idx + 1)

        if not sheet_created:
            ws = wb.active
            ws.title = sheet_name
            sheet_created = True
        else:
            existing = [s.title for s in wb.worksheets]
            if sheet_name in existing:
                sheet_name = "{} {}".format(sheet_name, t_idx + 1)
            ws = wb.create_sheet(sheet_name)

        ws.sheet_view.showGridLines = False
        current_row = 1

        # ── Title row ─────────────────────────────────────────
        _merge_row(ws, current_row, num_cols,
                   content["doc_title"],
                   bold=True, size=12,
                   bg=S_TITLE_BG, fg=S_WHITE_FG, height=28)
        current_row += 1

        # ── Subtitle row (deduplication check) ────────────────
        if (content["doc_subtitle"]
                and content["doc_subtitle"] not in content["doc_title"]
                and content["doc_title"] not in content["doc_subtitle"]
                and len(content["doc_subtitle"]) > 10):
            _merge_row(ws, current_row, num_cols,
                       content["doc_subtitle"],
                       bold=False, size=10,
                       bg=S_SUBHEAD_BG, fg=S_WHITE_FG, height=20)
            current_row += 1

        # ── Section heading row ───────────────────────────────
        if heading and heading != content["doc_title"]:
            _merge_row(ws, current_row, num_cols,
                       heading,
                       bold=True, size=10,
                       bg=S_SUBHEAD_BG, fg=S_WHITE_FG, height=20)
            current_row += 1

        # ── Column headers ────────────────────────────────────
        for ci, h in enumerate(headers, 1):
            _wc(ws, current_row, ci, h,
                bold=True, size=10,
                bg=S_HEADER_BG, fg=S_DARK_FG,
                align="center")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # ── Data rows ─────────────────────────────────────────
        col_widths = [len(str(h)) for h in headers]

        for ri, row in enumerate(rows):
            bg = S_ALT1_BG if ri % 2 == 0 else S_ALT2_BG
            for ci, val in enumerate(row, 1):
                col_widths[ci-1] = max(
                    col_widths[ci-1], len(str(val)) if val else 0)
                is_num = bool(re.match(
                    r"^\s*[\d,]+\.?\d*\s*(ML|ml|L|BEER|BOTTEL)?\s*$",
                    str(val)
                ))
                align = "right" if is_num else \
                        ("center" if ci == 1 else "left")
                _wc(ws, current_row, ci, val,
                    size=10, bg=bg, fg=S_BLACK_FG, align=align)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        # ── Total row ─────────────────────────────────────────
        total_label = "Total Records: {}".format(len(rows))
        ws.merge_cells("A{}:{}{}".format(
            current_row,
            get_column_letter(max(num_cols - 1, 1)),
            current_row
        ))
        _wc(ws, current_row, 1, total_label,
            bold=True, size=10,
            bg=S_TOTAL_BG, fg=S_DARK_FG, align="right")
        _wc(ws, current_row, num_cols, len(rows),
            bold=True, size=10,
            bg=S_TOTAL_BG, fg=S_DARK_FG, align="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # ── Apply column widths ────────────────────────────────
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                get_column_letter(ci)
            ].width = min(max(w + 4, 10), 60)

    # ── B: Write metadata/key-value pairs if present ─────────
    if content["metadata_pairs"]:

        sheet_name = "Details"
        if not sheet_created:
            ws_m = wb.active
            ws_m.title = sheet_name
            sheet_created = True
        else:
            ws_m = wb.create_sheet(sheet_name)

        ws_m.sheet_view.showGridLines = False
        r = 1

        ws_m.merge_cells("A{}:B{}".format(r, r))
        _wc(ws_m, r, 1, content["doc_title"],
            bold=True, size=12,
            bg=S_TITLE_BG, fg=S_WHITE_FG,
            align="center")
        ws_m.row_dimensions[r].height = 28
        r += 1

        # Subtitle (with deduplication check)
        if (content["doc_subtitle"]
                and content["doc_subtitle"] not in content["doc_title"]
                and content["doc_title"] not in content["doc_subtitle"]
                and len(content["doc_subtitle"]) > 10):
            ws_m.merge_cells("A{}:B{}".format(r, r))
            _wc(ws_m, r, 1, content["doc_subtitle"],
                bold=False, size=10,
                bg=S_SUBHEAD_BG, fg=S_WHITE_FG,
                align="center")
            ws_m.row_dimensions[r].height = 20
            r += 1

        # Column headers — "Field" | "Value"
        _wc(ws_m, r, 1, "Field",
            bold=True, size=10,
            bg=S_HEADER_BG, fg=S_DARK_FG, align="center")
        _wc(ws_m, r, 2, "Value",
            bold=True, size=10,
            bg=S_HEADER_BG, fg=S_DARK_FG, align="center")
        ws_m.row_dimensions[r].height = 22
        r += 1

        # Key-Value rows
        for ri, (key, val) in enumerate(
                content["metadata_pairs"]):
            bg = S_ALT1_BG if ri % 2 == 0 else S_ALT2_BG
            _wc(ws_m, r, 1, key,
                bold=True, size=10,
                bg=S_KEY_BG, fg=S_DARK_FG, align="left")
            _wc(ws_m, r, 2, val,
                size=10, bg=bg, fg=S_BLACK_FG, align="left")
            ws_m.row_dimensions[r].height = 18
            r += 1

        # Column widths
        key_vals   = [k for k, v in content["metadata_pairs"]]
        value_vals = [v for k, v in content["metadata_pairs"]]
        ws_m.column_dimensions["A"].width = \
            _col_width(key_vals, "Field")
        ws_m.column_dimensions["B"].width = \
            _col_width(value_vals, "Value")

    # ── C: Fallback if nothing extracted ─────────────────────
    if not sheet_created:
        ws = wb.active
        ws.title = "Extracted"
        ws.merge_cells("A1:D1")
        _wc(ws, 1, 1,
            "No structured data could be extracted from this PDF.",
            bold=True, size=11,
            bg=S_TITLE_BG, fg=S_WHITE_FG, align="center")

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════
# MAIN ENTRY POINT — WITH DOCUMENT TYPE DETECTION
# ═══════════════════════════════════════════════════════════

def extract_and_build(pdf_path: str, output_path: str) -> str:
    """
    Detects document type first, then uses the correct
    extraction strategy. Builds clean formatted Excel.

    - DOC_MIXED_BILL  → extract_bill_content()
    - DOC_PURE_TABLE  → extract_pdf_content()
    - DOC_FORM        → extract_pdf_content()
    """
    doc_type = detect_document_type(pdf_path)

    if doc_type == DOC_MIXED_BILL:
        content = extract_bill_content(pdf_path)
    else:
        content = extract_pdf_content(pdf_path)

    return build_excel(content, output_path)


# ═══════════════════════════════════════════════════════════
# JSON SUMMARY (for route.js process handler DB storage)
# ═══════════════════════════════════════════════════════════

def get_extraction_summary(pdf_path: str) -> dict:
    """
    Extract PDF, generate xlsx, and return a JSON-friendly summary.
    Called by scripts/extract.py which outputs JSON to stdout.
    """
    output_dir = os.path.join(tempfile.gettempdir(), "love2excel_outputs")
    os.makedirs(output_dir, exist_ok=True)
    xlsx_id = str(uuid.uuid4())
    xlsx_path = os.path.join(output_dir, "{}.xlsx".format(xlsx_id))

    try:
        # Use type-aware extraction
        doc_type = detect_document_type(pdf_path)

        if doc_type == DOC_MIXED_BILL:
            content = extract_bill_content(pdf_path)
        else:
            content = extract_pdf_content(pdf_path)

        build_excel(content, xlsx_path)

        tables   = content["tables"]
        metadata = content["metadata_pairs"]

        if tables:
            primary = max(tables, key=lambda t: len(t["rows"]))
            columns = primary["headers"]
            rows = []
            for row_data in primary["rows"]:
                row_dict = {}
                for ci, col in enumerate(columns):
                    row_dict[col] = row_data[ci] if ci < len(row_data) else ""
                rows.append(row_dict)

            detected_type = doc_type
            if metadata:
                detected_type = "mixed"

            return {
                "columns": columns,
                "rows": rows,
                "document_type": detected_type,
                "confidence": 0.90,
                "extraction_method": "pdfplumber_universal_v4",
                "xlsx_path": xlsx_path,
                "doc_title": content["doc_title"],
                "total_tables": len(tables),
                "total_metadata_fields": len(metadata),
            }

        elif metadata:
            columns = ["Field", "Value"]
            rows = [{"Field": k, "Value": v} for k, v in metadata]
            return {
                "columns": columns,
                "rows": rows,
                "document_type": "form",
                "confidence": 0.80,
                "extraction_method": "pdfplumber_universal_v4",
                "xlsx_path": xlsx_path,
                "doc_title": content["doc_title"],
                "total_tables": 0,
                "total_metadata_fields": len(metadata),
            }

        else:
            return {
                "columns": [],
                "rows": [],
                "document_type": "unknown",
                "confidence": 0.0,
                "error": "No structured data found in PDF",
                "extraction_method": "pdfplumber_universal_v4",
                "xlsx_path": "",
                "doc_title": content.get("doc_title", ""),
            }

    except Exception as exc:
        if os.path.exists(xlsx_path):
            os.remove(xlsx_path)
        return {
            "columns": [],
            "rows": [],
            "document_type": "unknown",
            "confidence": 0.0,
            "error": str(exc),
            "extraction_method": "pdfplumber_universal_v4",
            "xlsx_path": "",
        }
